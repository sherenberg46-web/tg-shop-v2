"""Admin panel API endpoints."""
import hashlib
import math
from io import BytesIO
from pathlib import Path
from typing import Optional

from pydantic import BaseModel

import aiosqlite
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Header, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill, Alignment

from backend.config import settings
from backend.database import get_db
from backend.schemas import ProductIn, AdminLoginIn, AdminStatusIn

router = APIRouter(prefix="/admin", tags=["admin"])

UPLOAD_DIR = Path("static/uploads")


def _expected_token() -> str:
    return hashlib.sha256(settings.ADMIN_PASSWORD.encode()).hexdigest()


def _require_admin(x_admin_token: Optional[str] = Header(None)) -> None:
    if x_admin_token != _expected_token():
        raise HTTPException(status_code=403, detail="Forbidden")


def _to_byn(price: float, region: str) -> float:
    """Mirror of frontend toBYN() from utils/price.ts."""
    if price <= 0:
        return 0
    if region in ("UA", "TR"):
        if price <= 500:
            f = 0.119
        elif price <= 1000:
            f = 0.112
        elif price <= 1500:
            f = 0.105
        else:
            f = 0.098
        return math.ceil(price * f)
    if region == "PL":
        if price <= 50:
            m = 1.70
        elif price <= 200:
            m = 1.60
        elif price <= 400:
            m = 1.50
        else:
            m = 1.40
        return math.ceil(price * 0.82 * m)
    if region == "IN":
        if price <= 500:
            m = 1.70
        elif price <= 1000:
            m = 1.60
        elif price <= 2000:
            m = 1.50
        else:
            m = 1.40
        return math.ceil(price * 0.040 * m)
    return math.ceil(price)


# ── Auth ──────────────────────────────────────────────────────────────────────

@router.post("/login")
async def admin_login(body: AdminLoginIn):
    if hashlib.sha256(body.password.encode()).hexdigest() != _expected_token():
        raise HTTPException(status_code=401, detail="Неверный пароль")
    return {"token": _expected_token()}


# ── Image Upload ──────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_image(
    file: UploadFile = File(...),
    _: None = Depends(_require_admin),
):
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    allowed = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Недопустимый формат файла")
    # Use a safe unique name
    import uuid
    filename = f"{uuid.uuid4().hex}{ext}"
    dest = UPLOAD_DIR / filename
    content = await file.read()
    dest.write_bytes(content)
    return {"url": f"/static/uploads/{filename}"}


# ── Products ──────────────────────────────────────────────────────────────────

@router.get("/products")
async def list_products(
    _: None = Depends(_require_admin),
    search: Optional[str] = None,
    category_id: Optional[int] = None,
    product_type: Optional[str] = None,
    platform: Optional[str] = None,
    is_active: Optional[int] = None,
    limit: int = Query(50, le=500),
    offset: int = 0,
    db: aiosqlite.Connection = Depends(get_db),
):
    where = "WHERE 1=1"
    params: list = []
    if search:
        where += " AND p.title LIKE ?"
        params.append(f"%{search}%")
    if category_id is not None:
        where += " AND p.category_id = ?"
        params.append(category_id)
    if product_type:
        where += " AND p.product_type = ?"
        params.append(product_type)
    if platform:
        where += " AND p.platform = ?"
        params.append(platform)
    if is_active is not None:
        where += " AND p.is_active = ?"
        params.append(is_active)

    async with db.execute(
        f"""SELECT p.id, p.category_id, p.title, p.description, p.image_url,
               p.price_uah, p.price_try, p.price_inr, p.price_pln,
               p.price_byn, p.price_byn_tr,
               p.platform, p.rating, p.is_featured, p.is_active,
               p.discount_pct, p.discount_until, p.product_type,
               p.created_at, p.updated_at,
               c.name AS category_name
            FROM products p
            LEFT JOIN categories c ON c.id = p.category_id
            {where}
            ORDER BY p.id DESC LIMIT ? OFFSET ?""",
        [*params, limit, offset],
    ) as cur:
        rows = await cur.fetchall()

    async with db.execute(
        f"SELECT COUNT(*) FROM products p {where}",
        params,
    ) as cur:
        total = (await cur.fetchone())[0]

    return {"items": [dict(r) for r in rows], "total": total}


@router.post("/products", status_code=201)
async def create_product(
    body: ProductIn,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    async with db.execute(
        """INSERT INTO products
           (category_id, title, description, image_url,
            price_uah, price_try, price_inr, price_pln,
            platform, rating, is_featured, is_active,
            discount_pct, discount_until,
            product_type, release_date, is_preorder)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            body.category_id, body.title, body.description, body.image_url,
            body.price_uah, body.price_try, body.price_inr, body.price_pln,
            body.platform, body.rating, int(body.is_featured), int(body.is_active),
            body.discount_pct, body.discount_until,
            body.product_type, body.release_date, int(body.is_preorder),
        ),
    ) as cur:
        product_id = cur.lastrowid
    await db.commit()
    return {"id": product_id, "message": "Товар создан"}


@router.put("/products/{product_id}")
async def update_product(
    product_id: int,
    body: ProductIn,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    await db.execute(
        """UPDATE products SET
           category_id=?, title=?, description=?, image_url=?,
           price_uah=?, price_try=?, price_inr=?, price_pln=?,
           price_byn=?, price_byn_tr=?,
           platform=?, rating=?, is_featured=?, is_active=?,
           discount_pct=?, discount_until=?,
           product_type=?, release_date=?, is_preorder=?,
           updated_at=CURRENT_TIMESTAMP
           WHERE id=?""",
        (
            body.category_id, body.title, body.description, body.image_url,
            body.price_uah, body.price_try, body.price_inr, body.price_pln,
            body.price_byn, body.price_byn_tr,
            body.platform, body.rating, int(body.is_featured), int(body.is_active),
            body.discount_pct, body.discount_until,
            body.product_type, body.release_date, int(body.is_preorder),
            product_id,
        ),
    )
    await db.commit()
    return {"message": "Товар обновлён"}


@router.delete("/products/{product_id}")
async def delete_product(
    product_id: int,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    await db.execute("DELETE FROM products WHERE id=?", (product_id,))
    await db.commit()
    return {"message": "Товар удалён"}


# ── Product Editions (admin-curated) ──────────────────────────────────────────

class EditionIn(BaseModel):
    name: str
    price_uah: Optional[float] = None
    price_try: Optional[float] = None
    price_inr: Optional[float] = None
    price_pln: Optional[float] = None
    is_default: bool = False
    sort_order: int = 0


@router.get("/products/{product_id}/editions")
async def get_product_editions_admin(
    product_id: int,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    rows = await db.execute_fetchall(
        "SELECT id, product_id, name, price_uah, price_try, price_inr, price_pln, is_default, sort_order FROM product_editions WHERE product_id=? ORDER BY sort_order",
        (product_id,),
    )
    return [dict(r) for r in rows]


@router.post("/products/{product_id}/editions", status_code=201)
async def add_product_edition(
    product_id: int,
    body: EditionIn,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    await db.execute(
        "INSERT INTO product_editions (product_id, name, price_uah, price_try, price_inr, price_pln, is_default, sort_order) VALUES (?,?,?,?,?,?,?,?)",
        (product_id, body.name, body.price_uah, body.price_try, body.price_inr, body.price_pln, int(body.is_default), body.sort_order),
    )
    await db.commit()
    return {"message": "Эдишен добавлен"}


@router.delete("/products/{product_id}/editions")
async def clear_product_editions(
    product_id: int,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    await db.execute("DELETE FROM product_editions WHERE product_id=?", (product_id,))
    await db.commit()
    return {"message": "Эдишены удалены"}


@router.delete("/products/{product_id}/editions/{edition_id}")
async def delete_product_edition(
    product_id: int,
    edition_id: int,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    await db.execute("DELETE FROM product_editions WHERE id=? AND product_id=?", (edition_id, product_id))
    await db.commit()
    return {"message": "Эдишен удалён"}


@router.post("/reset-discounts")
async def reset_discounts(
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    """Set discount_pct=0 and discount_until=NULL for ALL products."""
    cur = await db.execute(
        "SELECT COUNT(*) FROM products WHERE discount_pct > 0"
    )
    row = await cur.fetchone()
    count = row[0] if row else 0

    await db.execute(
        "UPDATE products SET discount_pct=0, discount_until=NULL, updated_at=CURRENT_TIMESTAMP WHERE discount_pct > 0"
    )
    await db.commit()
    return {"message": f"Сброшено скидок: {count}", "affected": count}


# ── Orders ────────────────────────────────────────────────────────────────────

def _order_where(date_from: Optional[str], date_to: Optional[str], status: Optional[str]):
    parts = ["1=1"]
    params: list = []
    if date_from:
        parts.append("DATE(o.created_at) >= ?")
        params.append(date_from)
    if date_to:
        parts.append("DATE(o.created_at) <= ?")
        params.append(date_to)
    if status:
        parts.append("o.status = ?")
        params.append(status)
    return "WHERE " + " AND ".join(parts), params


_ORDERS_SQL = """
    SELECT
        o.id, o.created_at, o.status, o.total_stars,
        u.username, u.first_name,
        (SELECT GROUP_CONCAT(p.title, ', ')
         FROM order_items oi2
         JOIN products p ON p.id = oi2.product_id
         WHERE oi2.order_id = o.id) AS items_title,
        (SELECT oi2.region FROM order_items oi2 WHERE oi2.order_id = o.id LIMIT 1) AS region,
        (SELECT SUM(oi2.price_paid) FROM order_items oi2 WHERE oi2.order_id = o.id) AS total_price_paid
    FROM orders o
    LEFT JOIN users u ON u.id = o.user_id
    {where}
    ORDER BY o.created_at DESC
"""


@router.get("/orders")
async def list_orders(
    _: None = Depends(_require_admin),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = Query(50, le=500),
    offset: int = 0,
    db: aiosqlite.Connection = Depends(get_db),
):
    where, params = _order_where(date_from, date_to, status)

    async with db.execute(
        _ORDERS_SQL.format(where=where) + " LIMIT ? OFFSET ?",
        [*params, limit, offset],
    ) as cur:
        rows = await cur.fetchall()

    async with db.execute(
        f"SELECT COUNT(*), COALESCE(SUM(total_stars), 0) FROM orders o {where}",
        params,
    ) as cur:
        agg = await cur.fetchone()

    orders = []
    for r in rows:
        d = dict(r)
        # price_paid in order_items already stores BYN values (converted on order creation)
        total_paid = d.get("total_price_paid") or 0
        d["total_byn"] = round(total_paid, 2)
        orders.append(d)

    total_stars = agg[1] or 0
    return {
        "items": orders,
        "total": agg[0] or 0,
        "total_stars": total_stars,
        "total_byn": round(total_stars / 23, 2),
    }


@router.get("/orders/export")
async def export_orders(
    _: None = Depends(_require_admin),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    db: aiosqlite.Connection = Depends(get_db),
):
    where, params = _order_where(date_from, date_to, status)

    async with db.execute(
        _ORDERS_SQL.format(where=where),
        params,
    ) as cur:
        rows = await cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "ID", "Дата", "Покупатель", "Username", "Товар(ы)",
        "Регион", "Сумма (нац.)", "Сумма (BYN)", "Звёзды", "Статус",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        d = dict(r)
        total_paid = d.get("total_price_paid") or 0
        total_byn = round(total_paid, 2)
        username = d.get("username") or ""
        ws.append([
            d["id"],
            d["created_at"],
            d.get("first_name") or "",
            f"@{username}" if username else "",
            d.get("items_title") or "",
            d.get("region") or "",
            round(total_paid, 2),
            total_byn,
            d["total_stars"],
            d["status"],
        ])

    col_widths = [6, 20, 18, 18, 40, 8, 14, 12, 10, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = f"{date_from or 'all'}_{date_to or 'all'}"
    filename = f"orders_{suffix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.put("/orders/{order_id}/status")
async def update_order_status(
    order_id: int,
    body: AdminStatusIn,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    valid = {"pending", "paid", "processing", "completed", "cancelled"}
    if body.status not in valid:
        raise HTTPException(status_code=400, detail=f"Допустимые статусы: {valid}")
    await db.execute(
        "UPDATE orders SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (body.status, order_id),
    )
    await db.commit()
    return {"message": "Статус обновлён"}

@router.get("/categories")
async def list_categories(_: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    async with db.execute("SELECT id, name, slug FROM categories ORDER BY name") as cur:
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@router.post("/categories")
async def create_category(body: dict, db: aiosqlite.Connection = Depends(get_db), _=Depends(_require_admin)):
    import re
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name обязателен")
    slug = body.get("slug", "").strip()
    if not slug:
        # Auto-generate slug from name
        slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or f"cat-{name[:20]}"
    await db.execute("INSERT INTO categories (name, slug) VALUES (?, ?)", (name, slug))
    await db.commit()
    async with db.execute("SELECT last_insert_rowid()") as cur:
        row = await cur.fetchone()
    return {"id": row[0], "name": name, "slug": slug}

@router.put("/categories/{cat_id}")
async def update_category(cat_id: int, body: dict, db: aiosqlite.Connection = Depends(get_db), _=Depends(_require_admin)):
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name обязателен")
    slug = body.get("slug", "").strip()
    if slug:
        await db.execute("UPDATE categories SET name=?, slug=? WHERE id=?", (name, slug, cat_id))
    else:
        await db.execute("UPDATE categories SET name=? WHERE id=?", (name, cat_id))
    await db.commit()
    return {"message": "Категория обновлена"}

@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: int, db: aiosqlite.Connection = Depends(get_db), _=Depends(_require_admin)):
    await db.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    await db.commit()
    return {"message": "Категория удалена"}


# ── Collections ───────────────────────────────────────────────────────────────

@router.get("/collections")
async def admin_list_collections(_: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    async with db.execute(
        """SELECT c.id, c.title, c.slug, c.description, c.is_active,
                  (SELECT COUNT(*) FROM collection_products WHERE collection_id = c.id) AS product_count
           FROM collections c ORDER BY c.title"""
    ) as cur:
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@router.post("/collections", status_code=201)
async def create_collection(body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    import re
    title = body.get("title", "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title обязателен")
    slug = body.get("slug", "").strip()
    if not slug:
        slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-") or f"col-{title[:20]}"
    desc = body.get("description", "")
    try:
        async with db.execute(
            "INSERT INTO collections(title, slug, description) VALUES(?,?,?)",
            (title, slug, desc)
        ) as cur:
            new_id = cur.lastrowid
        await db.commit()
        return {"id": new_id, "title": title, "slug": slug}
    except Exception:
        raise HTTPException(status_code=400, detail="Slug уже занят")

@router.put("/collections/{col_id}")
async def update_collection(col_id: int, body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    title = body.get("title", "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title обязателен")
    slug = body.get("slug", "").strip()
    desc = body.get("description", "")
    is_active = int(body.get("is_active", 1))
    if slug:
        await db.execute(
            "UPDATE collections SET title=?, slug=?, description=?, is_active=? WHERE id=?",
            (title, slug, desc, is_active, col_id)
        )
    else:
        await db.execute(
            "UPDATE collections SET title=?, description=?, is_active=? WHERE id=?",
            (title, desc, is_active, col_id)
        )
    await db.commit()
    return {"message": "Коллекция обновлена"}

@router.delete("/collections/{col_id}")
async def delete_collection(col_id: int, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    await db.execute("DELETE FROM collections WHERE id=?", (col_id,))
    await db.commit()
    return {"message": "Коллекция удалена"}

@router.get("/collections/{col_id}/products")
async def list_collection_products(col_id: int, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    async with db.execute(
        """SELECT p.id, p.title, p.image_url, p.price_uah, p.price_try, p.product_type, p.platform,
                  cp.sort_order
           FROM collection_products cp
           JOIN products p ON p.id = cp.product_id
           WHERE cp.collection_id = ?
           ORDER BY cp.sort_order, p.id""",
        (col_id,),
    ) as cur:
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@router.post("/collections/{col_id}/products")
async def add_collection_products(col_id: int, body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    product_ids = body.get("product_ids", [])
    if not product_ids:
        raise HTTPException(status_code=400, detail="product_ids обязателен")
    # Get current max sort_order
    async with db.execute(
        "SELECT COALESCE(MAX(sort_order), -1) FROM collection_products WHERE collection_id=?", (col_id,)
    ) as cur:
        row = await cur.fetchone()
    order = (row[0] + 1) if row else 0
    added = 0
    for pid in product_ids:
        try:
            await db.execute(
                "INSERT OR IGNORE INTO collection_products(collection_id, product_id, sort_order) VALUES(?,?,?)",
                (col_id, pid, order)
            )
            order += 1
            added += 1
        except Exception:
            pass
    await db.commit()
    return {"message": f"Добавлено товаров: {added}"}

@router.delete("/collections/{col_id}/products/{product_id}")
async def remove_collection_product(col_id: int, product_id: int, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    await db.execute(
        "DELETE FROM collection_products WHERE collection_id=? AND product_id=?",
        (col_id, product_id)
    )
    await db.commit()
    return {"message": "Товар убран из коллекции"}

@router.put("/collections/{col_id}/products/{product_id}/order")
async def set_product_order(col_id: int, product_id: int, body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    order = int(body.get("sort_order", 0))
    await db.execute(
        "UPDATE collection_products SET sort_order=? WHERE collection_id=? AND product_id=?",
        (order, col_id, product_id)
    )
    await db.commit()
    return {"message": "Порядок обновлён"}


# ── Statistics ────────────────────────────────────────────────────────────────

@router.get("/stats")
async def get_stats(
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    async with db.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN DATE(created_at) = DATE('now') THEN 1 ELSE 0 END), 0)          AS today_count,
            COALESCE(SUM(CASE WHEN created_at >= DATE('now', '-6 days') THEN 1 ELSE 0 END), 0)     AS week_count,
            COALESCE(SUM(CASE WHEN created_at >= DATE('now', '-29 days') THEN 1 ELSE 0 END), 0)    AS month_count,
            COALESCE(SUM(CASE WHEN DATE(created_at) = DATE('now') THEN total_stars ELSE 0 END), 0) AS today_stars,
            COALESCE(SUM(CASE WHEN created_at >= DATE('now', '-6 days') THEN total_stars ELSE 0 END), 0) AS week_stars,
            COALESCE(SUM(CASE WHEN created_at >= DATE('now', '-29 days') THEN total_stars ELSE 0 END), 0) AS month_stars
        FROM orders WHERE status != 'cancelled'
    """) as cur:
        agg = dict(await cur.fetchone())

    async with db.execute("""
        SELECT p.title, SUM(oi.quantity) AS cnt,
               SUM(oi.price_paid) AS revenue_local, oi.region
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        JOIN orders o ON o.id = oi.order_id
        WHERE o.status != 'cancelled'
          AND o.created_at >= DATE('now', '-29 days')
        GROUP BY oi.product_id, oi.region
        ORDER BY cnt DESC LIMIT 10
    """) as cur:
        top_rows = await cur.fetchall()

    async with db.execute("""
        SELECT oi.region,
               COUNT(DISTINCT o.id)  AS orders_count,
               SUM(oi.price_paid)    AS revenue_local
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        WHERE o.status != 'cancelled'
          AND o.created_at >= DATE('now', '-29 days')
        GROUP BY oi.region
    """) as cur:
        region_rows = await cur.fetchall()

    top_products = []
    for r in top_rows:
        d = dict(r)
        d["revenue_byn"] = _to_byn(d.get("revenue_local") or 0, d.get("region") or "UA")
        top_products.append(d)

    by_region = []
    for r in region_rows:
        d = dict(r)
        d["revenue_byn"] = _to_byn(d.get("revenue_local") or 0, d.get("region") or "UA")
        by_region.append(d)

    def stars_to_byn(s: int) -> int:
        return round(s / 23)

    return {
        "orders": {
            "today": agg["today_count"],
            "week": agg["week_count"],
            "month": agg["month_count"],
        },
        "revenue_byn": {
            "today": stars_to_byn(agg["today_stars"]),
            "week": stars_to_byn(agg["week_stars"]),
            "month": stars_to_byn(agg["month_stars"]),
        },
        "revenue_stars": {
            "today": agg["today_stars"],
            "week": agg["week_stars"],
            "month": agg["month_stars"],
        },
        "top_products": top_products,
        "by_region": by_region,
    }


# ── Settings (exchange rates) ─────────────────────────────────────────────────

DEFAULT_SETTINGS = {
    # Exchange rates
    "uah_rate":        "0.069",
    "try_rate":        "0.07",
    # UAH markup tiers (%)  — price ≤ threshold
    "uah_markup_500":  "70",
    "uah_markup_1000": "60",
    "uah_markup_1500": "50",
    "uah_markup_2000": "40",
    "uah_markup_max":  "30",
    # TRY markup tiers (%)
    "try_markup_500":  "70",
    "try_markup_1000": "60",
    "try_markup_1500": "50",
    "try_markup_2000": "40",
    "try_markup_max":  "30",
}

@router.get("/settings")
async def get_settings(_: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    async with db.execute("SELECT key, value FROM admin_settings") as cur:
        rows = await cur.fetchall()
    result = dict(DEFAULT_SETTINGS)
    for r in rows:
        result[r["key"]] = r["value"]
    return result

@router.put("/settings")
async def save_settings(body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    for key, value in body.items():
        await db.execute(
            "INSERT INTO admin_settings(key, value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, str(value))
        )
    await db.commit()
    return {"message": "Настройки сохранены"}


# ── Banners ───────────────────────────────────────────────────────────────────

@router.get("/banners")
async def list_banners(_: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    async with db.execute("SELECT * FROM admin_banners ORDER BY sort_order ASC, id ASC") as cur:
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@router.post("/banners", status_code=201)
async def create_banner(body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    await db.execute(
        """INSERT INTO admin_banners
           (title, subtitle, image_url, link_url, link_ua, link_tr, gradient, sort_order, is_active, collection_slug)
           VALUES(?,?,?,?,?,?,?,?,?,?)""",
        (body.get("title",""), body.get("subtitle",""), body.get("image_url",""),
         body.get("link_url",""), body.get("link_ua",""), body.get("link_tr",""),
         body.get("gradient",""), body.get("sort_order", 0), int(body.get("is_active", 1)),
         body.get("collection_slug",""))
    )
    await db.commit()
    return {"message": "Баннер создан"}

@router.put("/banners/{banner_id}")
async def update_banner(banner_id: int, body: dict, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    await db.execute(
        """UPDATE admin_banners
           SET title=?,subtitle=?,image_url=?,link_url=?,link_ua=?,link_tr=?,gradient=?,sort_order=?,is_active=?,collection_slug=?
           WHERE id=?""",
        (body.get("title",""), body.get("subtitle",""), body.get("image_url",""),
         body.get("link_url",""), body.get("link_ua",""), body.get("link_tr",""),
         body.get("gradient",""), body.get("sort_order", 0), int(body.get("is_active", 1)),
         body.get("collection_slug",""), banner_id)
    )
    await db.commit()
    return {"message": "Баннер обновлён"}

@router.delete("/banners/{banner_id}")
async def delete_banner(banner_id: int, _: None = Depends(_require_admin), db: aiosqlite.Connection = Depends(get_db)):
    await db.execute("DELETE FROM admin_banners WHERE id=?", (banner_id,))
    await db.commit()
    return {"message": "Баннер удалён"}


# ── Import products from local parser ────────────────────────────────────────

class ImportProductItem(BaseModel):
    title: str
    description: Optional[str] = None
    image_url: Optional[str] = None
    price_uah: Optional[float] = None
    price_try: Optional[float] = None
    price_byn: Optional[int] = None
    price_byn_tr: Optional[int] = None
    platform: Optional[str] = "PS4/PS5"
    discount_pct: float = 0
    discount_until: Optional[str] = None
    original_id: Optional[str] = None
    release_date: Optional[str] = None
    product_type: str = "game"
    is_preorder: bool = False
    is_featured: bool = False
    rating: float = 0
    region: str = "UA"
    task_type: Optional[str] = ""
    genre: Optional[str] = None
    requires_ps_plus: bool = False


class ImportBody(BaseModel):
    products: list[ImportProductItem]


@router.post("/import-products")
async def import_products(
    body: ImportBody,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    """
    Bulk upsert products from the local parser into this DB.
    Matches by original_id (preferred) then title+region.
    Returns counts of inserted and updated products.
    """
    from backend.routers.products import _load_price_settings, _uah_to_byn, _try_to_byn
    import math as _math

    ps = await _load_price_settings(db)
    inserted = updated = 0
    errors_detail: list[str] = []

    for item in body.products:
        try:
            title        = item.title
            original_id  = item.original_id or None  # keep NULL to avoid UNIQUE conflicts
            region       = item.region.upper()

            # Compute BYN prices if not provided
            price_byn = item.price_byn
            price_byn_tr = item.price_byn_tr
            if price_byn is None:
                if item.price_uah:
                    raw = _uah_to_byn(item.price_uah, ps)
                    price_byn = max(raw, 10) if raw > 0 else None
                elif item.price_try:
                    # TRY-only products (e.g. TR region): compute price_byn as fallback
                    raw = _try_to_byn(item.price_try, ps)
                    price_byn = max(raw, 10) if raw > 0 else None
            if price_byn_tr is None and item.price_try:
                raw = _try_to_byn(item.price_try, ps)
                price_byn_tr = max(raw, 10) if raw > 0 else None

            # Find existing
            product_id = None
            if original_id:
                async with db.execute(
                    "SELECT id FROM products WHERE original_id=?", (original_id,)
                ) as cur:
                    row = await cur.fetchone()
                    if row:
                        product_id = row[0]
            if product_id is None:
                async with db.execute(
                    "SELECT id FROM products WHERE title=? AND region=?", (title, region)
                ) as cur:
                    row = await cur.fetchone()
                    if row:
                        product_id = row[0]

            if product_id is not None:
                await db.execute(
                    """UPDATE products SET
                         title=?, description=?,
                         image_url=COALESCE(NULLIF(?, ''), image_url),
                         price_uah=COALESCE(?, price_uah),
                         price_try=COALESCE(?, price_try),
                         price_byn=COALESCE(?, price_byn),
                         price_byn_tr=COALESCE(?, price_byn_tr),
                         platform=COALESCE(NULLIF(?, ''), platform),
                         discount_pct=?, discount_until=COALESCE(?, discount_until),
                         release_date=COALESCE(?, release_date),
                         product_type=?, is_preorder=?,
                         is_featured=COALESCE(?, is_featured),
                         task_type=COALESCE(NULLIF(?, ''), task_type),
                         genre=COALESCE(NULLIF(?, ''), genre),
                         requires_ps_plus=?,
                         original_id=COALESCE(original_id, NULLIF(?, '')),
                         is_active=1, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (
                        title, item.description,
                        item.image_url or "",
                        item.price_uah, item.price_try, price_byn, price_byn_tr,
                        item.platform or "",
                        item.discount_pct, item.discount_until,
                        item.release_date,
                        item.product_type, int(item.is_preorder),
                        int(item.is_featured),
                        item.task_type or "",
                        item.genre or "",
                        int(item.requires_ps_plus),
                        original_id or "",
                        product_id,
                    ),
                )
                updated += 1
            else:
                # Ensure categories exist
                async with db.execute(
                    "SELECT id FROM categories LIMIT 1"
                ) as cur:
                    cat_row = await cur.fetchone()
                category_id = cat_row[0] if cat_row else 1

                await db.execute(
                    """INSERT INTO products
                         (category_id, title, description, image_url,
                          price_uah, price_try, price_byn, price_byn_tr,
                          platform, rating, is_featured, is_active,
                          discount_pct, discount_until, original_id, release_date,
                          product_type, is_preorder, region, task_type,
                          genre, requires_ps_plus)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        category_id, title, item.description, item.image_url,
                        item.price_uah, item.price_try, price_byn, price_byn_tr,
                        item.platform, item.rating, int(item.is_featured),
                        item.discount_pct, item.discount_until,
                        original_id, item.release_date,
                        item.product_type, int(item.is_preorder),
                        region, item.task_type or "",
                        item.genre or "", int(item.requires_ps_plus),
                    ),
                )
                inserted += 1
        except Exception as _e:
            errors_detail.append(f"{item.title}: {_e}")

    await db.commit()
    return {"inserted": inserted, "updated": updated, "total": len(body.products), "errors": errors_detail}


# ── Import game editions ───────────────────────────────────────────────────────

class EditionItem(BaseModel):
    product_original_id: str
    edition_name: str
    price_uah: Optional[float] = None
    price_try: Optional[float] = None
    price_byn: Optional[float] = None
    price_byn_tr: Optional[float] = None
    old_price_uah: Optional[float] = None
    old_price_try: Optional[float] = None
    discount_pct: int = 0
    platform: Optional[str] = None
    ps_store_url: Optional[str] = ""
    region: str = "UA"
    is_free: int = 0


class EditionsBody(BaseModel):
    editions: list[EditionItem]


@router.post("/import-editions")
async def import_editions(
    body: EditionsBody,
    _: None = Depends(_require_admin),
    db: aiosqlite.Connection = Depends(get_db),
):
    """
    Bulk upsert game_editions from local parser to this DB.
    Matches parent product by original_id.
    """
    inserted = updated = skipped = 0

    for item in body.editions:
        # Find product by original_id
        async with db.execute(
            "SELECT id FROM products WHERE original_id=?", (item.product_original_id,)
        ) as cur:
            row = await cur.fetchone()
        if not row:
            skipped += 1
            continue
        product_id = row[0]

        # Check if edition already exists
        async with db.execute(
            "SELECT id FROM game_editions WHERE parent_product_id=? AND edition_name=? AND region=?",
            (product_id, item.edition_name, item.region),
        ) as cur:
            existing = await cur.fetchone()

        if existing:
            await db.execute(
                """UPDATE game_editions SET
                     price_uah=?, price_try=?, price_byn=?, price_byn_tr=?,
                     old_price_uah=?, old_price_try=?,
                     discount_pct=?, platform=?, ps_store_url=?, is_free=?,
                     updated_at=CURRENT_TIMESTAMP
                   WHERE id=?""",
                (item.price_uah, item.price_try, item.price_byn, item.price_byn_tr,
                 item.old_price_uah, item.old_price_try,
                 item.discount_pct, item.platform, item.ps_store_url, item.is_free,
                 existing[0]),
            )
            updated += 1
        else:
            await db.execute(
                """INSERT INTO game_editions
                     (parent_product_id, edition_name,
                      price_uah, price_try, price_byn, price_byn_tr,
                      old_price_uah, old_price_try,
                      discount_pct, platform, ps_store_url, region, is_free, is_active)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
                (product_id, item.edition_name,
                 item.price_uah, item.price_try, item.price_byn, item.price_byn_tr,
                 item.old_price_uah, item.old_price_try,
                 item.discount_pct, item.platform, item.ps_store_url, item.region, item.is_free),
            )
            inserted += 1

    # Update has_editions flag for affected products
    await db.execute(
        """UPDATE products SET has_editions=1 WHERE original_id IN (
             SELECT DISTINCT p.original_id FROM game_editions ge
             JOIN products p ON p.id = ge.parent_product_id
           )"""
    )
    await db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}
